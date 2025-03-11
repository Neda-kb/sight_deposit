
import openpyxl as pyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
import hcob


def template_excel(df,outputpath,sheet_name, merge_cells=False):
    wbo = pyxl.Workbook()
    sheet = wbo.create_sheet(sheet_name)

    #Remove the default sheet
    del wbo['Sheet']

    #Iterate over the DataFrame and write to the sheet
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = sheet.cell(row=r_idx, column=c_idx, value=value)
            # Apply styles if on the header row
            if r_idx == 1:
                cell.style = hcob.HCOB_STYLES["HL1"]


    # Freeze the first row
    sheet.freeze_panes = "A2"

    if merge_cells:

        pass


    wbo.save(outputpath)
    return


def output(df_sight):
    """
    Aggregate the input DataFrame `df_sight` by specified columns and sum relevant numerical columns.

    """
    cols_to_sum = ["NII", "interest_bearing amount"]
    cols_to_group = ["", "", "g", ""]

    df_sight_output = df_sight[cols_to_sum + cols_to_group].groupby(cols_to_group).sum().reset_index()


    df_sight_output[""] = df_sight_output["NII"] / df_sight_output["interest_bearing_amount"]

    # Add a new column '' based on the value of ''
    currency_mapping = {"EUR": "EUR", "USD": "USD"}
    df_sight_output[""] = df_sight_output[""].map(currency_mapping).fillna("")

    # Create a new column 'CLUSTER' by concatenating 'OIPS' and 'CurrencyAGG' with a space in between
    df_sight_output[""] = df_sight_output[""] + " " + df_sight_output[""]

    return df_sight_output

